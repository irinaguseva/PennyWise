import os

import openpyxl
from openpyxl.styles import Font

from .CONF import generate_report_filename


def generate_test_excel(report_data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    ws["A1"] = "Financial Report"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A2"] = (
        f"Period: {report_data['period']['start']} to {report_data['period']['end']}"
    )

    headers = ["Category", "Income", "Expense", "Balance"]
    ws.append(headers)

    for category, values in report_data["categories"].items():
        ws.append(
            [
                category,
                values["income"],
                values["expense"],
                values["income"] - values["expense"],
            ]
        )

    ws.append(
        [
            "TOTAL",
            report_data["total"]["income"],
            report_data["total"]["expense"],
            report_data["total"]["balance"],
        ]
    )

    for cell in ws[3]:
        cell.font = Font(bold=True)

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    output_dir = "reports"
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, generate_report_filename("report_table"))
    wb.save(filepath)
    print(f"Файл успешно сохранен: {os.path.abspath(filepath)}")
    return filepath
