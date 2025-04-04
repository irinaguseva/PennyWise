# budget/excel_utils.py
import openpyxl
from openpyxl.styles import Font
from io import BytesIO


def generate_category_report_excel(report_data):
    """Генерация Excel-отчёта"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Financial Report"

    # Заголовок
    ws['A1'] = "Financial Report"
    ws['A1'].font = Font(bold=True, size=14)

    # Период отчёта
    ws['A2'] = f"Period: {report_data['period']['start']} to {report_data['period']['end']}"

    # Шапка таблицы
    headers = ["Category", "Income", "Expense", "Balance"]
    ws.append(headers)

    # Данные
    for category, values in report_data['categories'].items():
        ws.append([
            category,
            values['income'],
            values['expense'],
            values['income'] - values['expense']
        ])

    # Итоги
    ws.append([
        "TOTAL",
        report_data['total']['income'],
        report_data['total']['expense'],
        report_data['total']['balance']
    ])

    # Форматирование
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Сохранение в буфер
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer